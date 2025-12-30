import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List
from .oracle_mapper import ColumnInfo, TableInfo


class ExcelWriter:
    """Excel 출력 모듈"""

    def __init__(self):
        self.wb = Workbook()
        # 기본 스타일
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        self.section_font = Font(bold=True, size=12)
        self.section_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.section_font_white = Font(bold=True, size=12, color="FFFFFF")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.unmapped_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    def create_description_sheet(self, proc_name: str, description: str, parameters: list):
        """설명 시트 생성"""
        ws = self.wb.active
        ws.title = "설명"

        current_row = 1

        # 프로시저 이름
        cell = ws.cell(row=current_row, column=1, value="프로시저 이름")
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.border = self.thin_border
        ws.cell(row=current_row, column=2, value=proc_name).border = self.thin_border
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        current_row += 2

        # 파라미터 섹션
        cell = ws.cell(row=current_row, column=1, value="[파라미터]")
        cell.font = self.section_font_white
        cell.fill = self.section_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1

        # 파라미터 헤더
        headers = ["파라미터명", "데이터타입"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        current_row += 1

        # 파라미터 데이터
        for param in parameters:
            if isinstance(param, dict):
                name = param.get("name", "")
                ptype = param.get("type", "")
            else:
                name = str(param)
                ptype = ""
            ws.cell(row=current_row, column=1, value=name).border = self.thin_border
            ws.cell(row=current_row, column=2, value=ptype).border = self.thin_border
            current_row += 1

        current_row += 1

        # 설명 섹션
        cell = ws.cell(row=current_row, column=1, value="[프로시저 설명]")
        cell.font = self.section_font_white
        cell.fill = self.section_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1

        # 설명 내용
        desc_cell = ws.cell(row=current_row, column=1, value=description)
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 5, end_column=4)
        ws.row_dimensions[current_row].height = 100

        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def create_sheet(self, title: str, table_infos: List[TableInfo], column_infos: List[ColumnInfo]):
        """시트 생성 (입력/출력용)"""
        # 설명 시트가 이미 있으면 새 시트 생성
        ws = self.wb.create_sheet(title=title)

        current_row = 1

        # === 테이블 정보 섹션 ===
        current_row = self._write_section_header(ws, current_row, "테이블 정보")
        current_row = self._write_table_info(ws, current_row, table_infos)
        current_row += 1  # 빈 줄

        # === 항목 정보 섹션 ===
        current_row = self._write_section_header(ws, current_row, "항목 정보")
        current_row = self._write_column_info(ws, current_row, column_infos)

        # 컬럼 너비 조정
        self._auto_adjust_column_width(ws)

    def _write_section_header(self, ws, row: int, title: str) -> int:
        """섹션 헤더 작성"""
        cell = ws.cell(row=row, column=1, value=f"[{title}]")
        cell.font = self.section_font_white
        cell.fill = self.section_fill
        # 병합
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        return row + 1

    def _write_table_info(self, ws, row: int, table_infos: List[TableInfo]) -> int:
        """테이블 정보 작성"""
        # 헤더
        headers = ["관련테이블 한글명", "관련테이블 영문명"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        row += 1

        # 데이터 (중복 제거)
        seen = set()
        for info in table_infos:
            key = (info.table_kor, info.table_eng)
            if key in seen:
                continue
            seen.add(key)

            cell1 = ws.cell(row=row, column=1, value=info.table_kor)
            cell2 = ws.cell(row=row, column=2, value=info.table_eng)

            cell1.border = self.thin_border
            cell2.border = self.thin_border

            # 매핑 없음 표시
            if not info.is_mapped:
                cell1.fill = self.unmapped_fill
                cell2.fill = self.unmapped_fill

            row += 1

        return row

    def _write_column_info(self, ws, row: int, column_infos: List[ColumnInfo]) -> int:
        """항목 정보 작성"""
        # 헤더
        headers = ["테이블 한글명", "테이블 영문명", "항목 한글명", "항목 영문명", "유형", "길이", "PK", "FK", "기존 테이블 영문명"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        row += 1

        # 데이터
        for info in column_infos:
            values = [
                info.table_kor,
                info.table_eng,
                info.col_kor,
                info.col_eng,
                info.data_type,
                info.length,
                info.pk,
                info.fk,
                info.old_table_eng
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.thin_border

                # 매핑 없음 표시
                if not info.is_mapped:
                    cell.fill = self.unmapped_fill

            row += 1

        return row

    def _auto_adjust_column_width(self, ws):
        """컬럼 너비 자동 조정"""
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils import get_column_letter

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    # 병합된 셀은 건너뛰기
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value:
                        # 한글은 2배 너비
                        length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                        max_length = max(max_length, length)
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def save(self, filepath: str):
        """파일 저장"""
        self.wb.save(filepath)
        print(f"Excel 파일 저장 완료: {filepath}")

    def save_csv(self, filepath: str, table_infos: List[TableInfo], column_infos: List[ColumnInfo], sheet_type: str = ""):
        """CSV 파일 저장"""
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # 테이블 정보 섹션
            writer.writerow([f"[{sheet_type} 테이블 정보]"])
            writer.writerow(["관련테이블 한글명", "관련테이블 영문명", "매핑여부"])

            seen = set()
            for info in table_infos:
                key = (info.table_kor, info.table_eng)
                if key in seen:
                    continue
                seen.add(key)
                writer.writerow([info.table_kor, info.table_eng, "O" if info.is_mapped else "X"])

            writer.writerow([])  # 빈 줄

            # 항목 정보 섹션
            writer.writerow([f"[{sheet_type} 항목 정보]"])
            writer.writerow(["테이블 한글명", "테이블 영문명", "항목 한글명", "항목 영문명", "유형", "길이", "PK", "FK", "기존 테이블 영문명", "매핑여부"])

            for info in column_infos:
                writer.writerow([
                    info.table_kor,
                    info.table_eng,
                    info.col_kor,
                    info.col_eng,
                    info.data_type,
                    info.length,
                    info.pk,
                    info.fk,
                    info.old_table_eng,
                    "O" if info.is_mapped else "X"
                ])

        print(f"CSV 파일 저장 완료: {filepath}")


def test_writer():
    """Excel 작성기 테스트"""
    writer = ExcelWriter()

    # 샘플 데이터
    table_infos = [
        TableInfo(table_kor="후드코드", table_eng="TC_HCM_HOOD_CD"),
        TableInfo(table_kor="[매핑없음] V_TEST", table_eng="V_TEST", is_mapped=False),
    ]

    column_infos = [
        ColumnInfo(
            table_kor="후드코드", table_eng="TC_HCM_HOOD_CD",
            col_kor="코드명", col_eng="CD_NM",
            data_type="VARCHAR", length="20", pk="Y", fk=""
        ),
        ColumnInfo(
            table_kor="[매핑없음] V_TEST", table_eng="V_TEST",
            col_kor="[매핑없음] TestCol", col_eng="TestCol",
            data_type="", length="", pk="", fk="", is_mapped=False
        ),
    ]

    writer.create_sheet("입력", table_infos, column_infos)
    writer.create_sheet("출력", table_infos, column_infos)
    writer.save("test_output.xlsx")


if __name__ == "__main__":
    test_writer()
