"""
BOGUN2018 프로시저 매핑 추출기
- C# 소스에서 UP_NBOGUN_* 프로시저 호출을 추출
- 폼/Biz 클래스별로 분류하여 CSV/Excel 출력
"""

import os
import re
import csv
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Dict, Set
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class ProcedureCall:
    """프로시저 호출 정보"""
    procedure_name: str
    file_path: str
    line_number: int
    call_pattern: str  # DataTable, DataReader, ExecuteNonQuery
    form_name: str = ""  # 추론된 폼 이름


@dataclass
class FormSummary:
    """폼별 요약 정보"""
    form_name: str
    procedure_count: int = 0
    procedures: List[str] = field(default_factory=list)


@dataclass
class ProcedureSummary:
    """프로시저별 요약 정보"""
    procedure_name: str
    call_count: int = 0
    calling_forms: List[str] = field(default_factory=list)


# Biz 클래스 → 폼 매핑 테이블
BIZ_TO_FORM_MAPPING = {
    "BizPlan.cs": "frmPlan",
    "BizVisit.cs": "frmTotalMenu",
    "BizSaupjaCard.cs": "frmSaupjaCard",
    "BizJaehaeja.cs": "frmJaehaeja",
    "BizContract.cs": "frmEstimation",
    "BizExcel.cs": "frmExcel",
    "BizSangdamPersonD1Jido.cs": "frmSangdamPersonD1Jido",
    "BizSaupjaDangerMaterial.cs": "frmDangerMaterial",
    "BizBonbu.cs": "frmBonbu",
    "BizBohogu.cs": "frmBohogu",
    "BizWeewonhoi.cs": "frmWeewonhoi",
    "BizKiupJisu.cs": "frmKiupJisu",
    "BizUpmuDamdang.cs": "frmUpmuDamdang",
    "BizSaupjaDangerMachine.cs": "frmDangerMachine",
    "BizJakupManage.cs": "frmJakupManage",
    "BizCode.cs": "CodeManagement",
    "BizSafeWorkplace.cs": "frmSafeWorkplace",
    "BizPreInwon.cs": "frmPreInwon",
    "BizJakupCheck.cs": "frmJakupCheck",
    "BizF6.cs": "frmF6Print",
    "BizSaupjaCheck.cs": "frmSaupjaCheck",
    "BizCommon.cs": "Common",
}

# 프로시저 이름 패턴 → 메뉴 매핑 (우선순위 순서대로 정렬)
# 더 구체적인 패턴이 먼저 오도록 정렬
PROC_NAME_PATTERNS = [
    # 상태보고서 (frmSaupjaDetail에서 사용 - 가장 구체적)
    # RPT_Sangtae, SaupjaUpmuDamdang, SaupjaUpmuManager 등
    (r"RPT_Sangtae|SaupjaUpmuDamdang|SaupjaUpmuManager|SaupjaDetail", "1-8. 보건관리 > 상태보고서"),

    # 사업장관리카드 (RPT_SaupjaCard - 가장 구체적)
    (r"RPT_SaupjaCard|SaupjaCard", "1-9. 보건관리 > 사업장관리카드"),

    # 작업환경 (JakupEnv 관련)
    (r"JakupEnv|VisitJakupEnv", "1-4. 보건관리 > 작업환경"),

    # 공정관리 (Jakup 하위 항목들 - Gongjung, Site, Position, Result, Sayong)
    (r"JakupGongjung|JakupPosition|JakupSite|JakupResult|JakupSayong|JakupMove|JakupMan|GongjungSaupja", "1-2. 보건관리 > 공정관리"),
    (r"Gongjung", "1-2. 보건관리 > 공정관리"),

    # 작업관리 (JakupManage - 위의 패턴에 안 맞는 Jakup)
    (r"JakupManage|JakupCheck", "1-5. 보건관리 > 작업관리"),

    # 일반관리 (신규 추가)
    (r"Ilban|IlbanSahu", "1-3. 보건관리 > 일반관리"),

    # 연간계획
    (r"Plan|Target|PlanEdu", "1-1. 보건관리 > 연간계획"),

    # 건강관리/지도
    (r"Jido|Sangdam|SangdamPerson", "1-6. 보건관리 > 건강관리"),

    # 교육관리
    (r"Edu|EduItem|EduManage", "1-7. 보건관리 > 교육관리"),

    # 보고서/인쇄 (RPT_SaupjaCard 제외) → 메인 보고서
    (r"RPT_(?!SaupjaCard)|Print|F6|Report", "13. 메인 > 보고서"),

    # 지정 (신규 추가)
    (r"Jijung|JijungSTAND", "3. 메인 > 지정"),

    # 방문관리
    (r"Visit(?!Jakup)", "92. 방문관리"),  # VisitJakup 제외

    # 계약
    (r"Estimation|Contract", "2. 메인 > 계약"),

    # 인력
    (r"Person|Damdang|Inwon", "4. 메인 > 인력"),

    # 장비
    (r"Device|Jangbi|Jaego", "5. 메인 > 장비"),

    # 청구 (신규 추가)
    (r"Cheonggu", "7. 메인 > 실적청구"),

    # 통계
    (r"STATISTIC|Statistic|Chart", "8. 메인 > 통계"),

    # 엑셀
    (r"Excel", "9. 메인 > 엑셀"),

    # 본부
    (r"Bonbu", "10. 메인 > 본부"),

    # 건강증진지수
    (r"KiupJisu|Kiup", "11. 메인 > 기업건강증진지수"),

    # 헬스키퍼
    (r"Health|HealthUp", "12. 메인 > 헬스키퍼"),

    # 위험/안전평가
    (r"Danger|Pyeongga|DrPyeongga|Safe", "91. 위험/안전평가"),

    # 재해자/측정
    (r"Jaehaeja|Chkjn|Measur", "93. 재해자/측정"),

    # 보건관리 메인 (Saupja - 마지막에 체크, SaupjaCard 제외)
    (r"Saupja(?!Card)|Month|Bohogu|Material|Selbi", "1. 메인 > 보건관리"),

    # 코드관리
    (r"Code(?!Material)|DHCode|MethodCode", "90. 공통/코드관리"),
]

# 메뉴 구조 (001.png + 002.png 기반)
MENU_STRUCTURE = {
    "1. 메인 > 보건관리": {
        "description": "사업장 관리 메인 (frmSaupja)",
        "forms": ["frmSaupja", "frmSaupjaDetail", "frmSaupjaSearch",
                  "frmSaupjaList", "frmSaupjaMonth", "frmSaupjaUpmuDamdang"],
        "keywords": ["Saupja", "사업장"],
    },
    "1-1. 보건관리 > 연간계획": {
        "description": "연간계획 (frmPlan)",
        "forms": ["frmPlan", "frmPlanTarget", "frmPlanEduItem", "frmPlanEduSilsi"],
        "keywords": ["Plan", "연간", "계획"],
    },
    "1-2. 보건관리 > 공정관리": {
        "description": "공정관리 (frmGongjungManage)",
        "forms": ["frmGongjungManage", "frmGongjung", "frmGongjungMove", "frmGJMoveToSite"],
        "keywords": ["Gongjung", "공정"],
    },
    "1-3. 보건관리 > 일반관리": {
        "description": "일반관리 (frmIlbanManage)",
        "forms": ["frmIlbanManage"],
        "keywords": ["Ilban", "일반"],
    },
    "1-4. 보건관리 > 작업환경": {
        "description": "작업환경 (frmJakupEnv)",
        "forms": ["frmJakupEnv", "frmJakupHistory"],
        "keywords": ["JakupEnv", "작업환경"],
    },
    "1-5. 보건관리 > 작업관리": {
        "description": "작업관리 (frmJakupManage)",
        "forms": ["frmJakupManage", "frmJakup"],
        "keywords": ["JakupManage", "작업관리"],
    },
    "1-6. 보건관리 > 건강관리": {
        "description": "건강관리/지도 (frmJido)",
        "forms": ["frmJido", "frmJidoSearch", "frmJidoDevice", "frmJidoInwon",
                  "frmJidoJaehaeja", "frmJidoJakupEnv", "frmJidoMaterial",
                  "frmJidoSangdamSilsi", "frmJidoUpmuManage", "frmSangdam",
                  "frmSangdamPerson", "frmSangdamPos"],
        "keywords": ["Jido", "지도", "Sangdam", "상담", "건강"],
    },
    "1-7. 보건관리 > 교육관리": {
        "description": "교육관리 (frmEduManage)",
        "forms": ["frmEduManage", "frmEduItem", "frmEduItemManage"],
        "keywords": ["Edu", "교육"],
    },
    "1-8. 보건관리 > 상태보고서": {
        "description": "상태보고서 팝업 (frmSaupja 상단 버튼 → frmSaupjaDetail)",
        "forms": ["frmSaupjaDetail"],
        "keywords": ["상태보고서", "SaupjaDetail"],
    },
    "1-9. 보건관리 > 사업장관리카드": {
        "description": "사업장관리카드 (frmSaupjaCard)",
        "forms": ["frmSaupjaCard"],
        "keywords": ["SaupjaCard", "관리카드"],
    },
    "2. 메인 > 계약": {
        "description": "계약/추정 관리 (frmEstimation)",
        "forms": ["frmEstimation", "frmEstimationSaupja", "frmEstimationUpmu"],
        "keywords": ["Estimation", "Contract", "계약", "추정"],
    },
    "3. 메인 > 지정": {
        "description": "지정 관리 (frmJijung)",
        "forms": ["frmJijung"],
        "keywords": ["Jijung", "지정"],
    },
    "4. 메인 > 인력": {
        "description": "인사 관리 (frmPerson)",
        "forms": ["frmPerson", "frmDamdang", "frmUpmuDamdang", "frmCenterKikwan"],
        "keywords": ["Person", "인사", "인력", "Damdang", "담당"],
    },
    "5. 메인 > 장비": {
        "description": "장비 관리 (frmDeviceMonth)",
        "forms": ["frmDevice", "frmDeviceMonth", "frmDeviceRepair", "frmJangbi", "frmJaego"],
        "keywords": ["Device", "장비", "Jangbi", "재고"],
    },
    "6. 메인 > 부서장": {
        "description": "부서장 관리 (frmBuseojang)",
        "forms": ["frmBuseojang"],
        "keywords": ["Buseojang", "부서장"],
    },
    "7. 메인 > 실적청구": {
        "description": "청구 관리 (frmCheonggu)",
        "forms": ["frmCheonggu"],
        "keywords": ["Cheonggu", "청구", "실적"],
    },
    "8. 메인 > 통계": {
        "description": "통계 (frmStatistic)",
        "forms": ["frmStatistic", "frmStatisticsSub1", "frmChart", "frmF8Statistics"],
        "keywords": ["Statistic", "통계", "Chart"],
    },
    "9. 메인 > 엑셀": {
        "description": "엑셀 출력 (frmExcel)",
        "forms": ["frmExcel"],
        "keywords": ["Excel", "엑셀"],
    },
    "10. 메인 > 본부": {
        "description": "본부 관리 (frmBonbu)",
        "forms": ["frmBonbu"],
        "keywords": ["Bonbu", "본부"],
    },
    "11. 메인 > 기업건강증진지수": {
        "description": "건강증진지수 (frmKiupJisu)",
        "forms": ["frmKiupJisu"],
        "keywords": ["KiupJisu", "건강증진", "지수"],
    },
    "12. 메인 > 헬스키퍼": {
        "description": "건강지킴이 (frmHealthKeeperInvite)",
        "forms": ["frmHealthKeeperInvite", "frmHealthUp"],
        "keywords": ["Health", "헬스"],
    },
    "13. 메인 > 보고서": {
        "description": "보고서 (Form1 → frmF6Print)",
        "forms": ["frmF6Print", "frmPrint", "frmExportLog"],
        "keywords": ["Print", "보고서", "출력", "F6", "RPT"],
    },
    "90. 공통/코드관리": {
        "description": "코드 관리",
        "forms": ["CodeManagement", "frmCode"],
        "keywords": ["Code", "코드"],
    },
    "91. 위험/안전평가": {
        "description": "위험평가 관련",
        "forms": ["frmDangerMaterial", "frmDangerPyeongga", "frmDrPyeonggaList"],
        "keywords": ["Danger", "위험", "Pyeongga", "평가", "안전"],
    },
    "92. 방문관리": {
        "description": "방문 통합메뉴",
        "forms": ["frmTotalMenu", "frmVisitList", "frmVisitReport", "frmVisitPosition"],
        "keywords": ["Visit", "방문", "TotalMenu"],
    },
    "93. 재해자/측정": {
        "description": "재해자 및 측정 관리",
        "forms": ["frmJaehaeja", "frmChkjn", "frmChkjnDate", "frmChkjnFavorite"],
        "keywords": ["Jaehaeja", "재해", "Chkjn", "측정"],
    },
    "99. 공통모듈": {
        "description": "공통 비즈니스 로직 (Biz.cs)",
        "forms": ["Common(Biz.cs)", "Common"],
        "keywords": [],
    },
}


class BogunProcedureExtractor:
    """BOGUN2018 프로시저 추출기"""

    def __init__(self, bogun_path: str):
        self.bogun_path = Path(bogun_path)
        self.procedure_calls: List[ProcedureCall] = []
        self.form_summaries: Dict[str, FormSummary] = {}
        self.procedure_summaries: Dict[str, ProcedureSummary] = {}

        # 정규식 패턴
        self.proc_pattern = re.compile(r'["\']UP_NBOGUN_(\w+)["\']')
        self.call_patterns = {
            'DataTable': re.compile(r'\.DataTable\s*\('),
            'DataSet': re.compile(r'\.DataSet\s*\('),
            'DataReader': re.compile(r'\.DataReader\s*\('),
            'ExecuteNonQuery': re.compile(r'\.ExecuteNonQuery\s*\('),
            'ExecuteScalar': re.compile(r'\.ExecuteScalar\s*\('),
        }

    def extract_all(self):
        """모든 .cs 파일에서 프로시저 추출"""
        print(f"스캔 시작: {self.bogun_path}")

        cs_files = list(self.bogun_path.rglob("*.cs"))
        print(f"발견된 C# 파일: {len(cs_files)}개")

        for cs_file in cs_files:
            self._extract_from_file(cs_file)

        print(f"추출 완료: {len(self.procedure_calls)}건")
        print(f"고유 프로시저: {len(self._get_unique_procedures())}개")

        # 요약 정보 생성
        self._build_summaries()

    def _extract_from_file(self, file_path: Path):
        """단일 파일에서 프로시저 추출"""
        try:
            # UTF-8 BOM 처리
            with open(file_path, 'r', encoding='utf-8-sig', errors='ignore') as f:
                lines = f.readlines()
        except Exception as e:
            print(f"파일 읽기 오류: {file_path} - {e}")
            return

        for line_num, line in enumerate(lines, 1):
            matches = self.proc_pattern.findall(line)
            if matches:
                for proc_name in matches:
                    full_proc_name = f"UP_NBOGUN_{proc_name}"
                    call_pattern = self._detect_call_pattern(line)
                    form_name = self._infer_form_name(file_path)

                    # 상대 경로 저장
                    rel_path = str(file_path.relative_to(self.bogun_path.parent))

                    self.procedure_calls.append(ProcedureCall(
                        procedure_name=full_proc_name,
                        file_path=rel_path,
                        line_number=line_num,
                        call_pattern=call_pattern,
                        form_name=form_name
                    ))

    def _detect_call_pattern(self, line: str) -> str:
        """호출 패턴 감지"""
        for pattern_name, pattern in self.call_patterns.items():
            if pattern.search(line):
                return pattern_name
        return "Unknown"

    def _infer_form_name(self, file_path: Path) -> str:
        """파일 경로에서 폼 이름 추론"""
        filename = file_path.name

        # frm*.cs 파일 → 직접 매핑
        if filename.startswith("frm") and filename.endswith(".cs"):
            return filename[:-3]  # .cs 제거

        # Biz*.cs 파일 → 매핑 테이블 사용
        if filename in BIZ_TO_FORM_MAPPING:
            return BIZ_TO_FORM_MAPPING[filename]

        # Biz.cs (메인 비즈니스 클래스) → Common
        if filename == "Biz.cs":
            return "Common(Biz.cs)"

        # UC_*.cs (커스텀 컨트롤) → 컨트롤 이름
        if filename.startswith("UC_") and filename.endswith(".cs"):
            return filename[:-3]

        # Sub 폴더 내 파일
        if "Sub" in str(file_path):
            return f"Sub/{filename[:-3]}"

        # ModalPopup 폴더 내 파일
        if "ModalPopup" in str(file_path):
            return f"ModalPopup/{filename[:-3]}"

        # Classes 폴더
        if "Classes" in str(file_path):
            return f"Classes/{filename[:-3]}"

        # CostomControl 폴더
        if "CostomControl" in str(file_path):
            return f"CustomControl/{filename[:-3]}"

        return f"Other/{filename[:-3]}"

    def _get_unique_procedures(self) -> Set[str]:
        """고유 프로시저 목록"""
        return set(call.procedure_name for call in self.procedure_calls)

    def _build_summaries(self):
        """요약 정보 생성"""
        # 폼별 요약
        form_procs = defaultdict(set)
        for call in self.procedure_calls:
            form_procs[call.form_name].add(call.procedure_name)

        for form_name, procs in form_procs.items():
            self.form_summaries[form_name] = FormSummary(
                form_name=form_name,
                procedure_count=len(procs),
                procedures=sorted(list(procs))
            )

        # 프로시저별 요약
        proc_forms = defaultdict(set)
        for call in self.procedure_calls:
            proc_forms[call.procedure_name].add(call.form_name)

        for proc_name, forms in proc_forms.items():
            # 호출 횟수 계산
            call_count = sum(1 for c in self.procedure_calls if c.procedure_name == proc_name)
            self.procedure_summaries[proc_name] = ProcedureSummary(
                procedure_name=proc_name,
                call_count=call_count,
                calling_forms=sorted(list(forms))
            )

    def _classify_by_menu(self, form_name: str, proc_name: str) -> str:
        """폼/프로시저를 메뉴 구조에 따라 분류"""
        # 0. 공통모듈(Biz.cs)에서 호출된 경우 → 프로시저 이름 패턴으로 우선 분류
        if "Biz.cs" in form_name or form_name == "Common":
            for pattern, menu_key in PROC_NAME_PATTERNS:
                if re.search(pattern, proc_name, re.IGNORECASE):
                    return menu_key
            # 패턴 매칭 실패 시 99. 공통모듈로 분류
            return "99. 공통모듈"

        # 1. 폼 이름으로 직접 매칭
        for menu_key, menu_info in MENU_STRUCTURE.items():
            if form_name in menu_info["forms"]:
                return menu_key

        # 2. 폼 이름 키워드 매칭
        for menu_key, menu_info in MENU_STRUCTURE.items():
            for keyword in menu_info["keywords"]:
                if keyword.lower() in form_name.lower():
                    return menu_key

        # 3. 프로시저 이름 패턴 매칭 (일반 폼에서 호출된 경우에도 적용)
        for pattern, menu_key in PROC_NAME_PATTERNS:
            if re.search(pattern, proc_name, re.IGNORECASE):
                return menu_key

        # 4. MENU_STRUCTURE 키워드 매칭 (fallback)
        for menu_key, menu_info in MENU_STRUCTURE.items():
            for keyword in menu_info["keywords"]:
                if keyword.lower() in proc_name.lower():
                    return menu_key

        return "98. 미분류"

    def get_menu_based_summary(self) -> Dict[str, Dict]:
        """메뉴 기반 요약 생성"""
        menu_summary = defaultdict(lambda: {
            "procedures": set(),
            "forms": set(),
            "call_count": 0
        })

        for call in self.procedure_calls:
            menu = self._classify_by_menu(call.form_name, call.procedure_name)
            menu_summary[menu]["procedures"].add(call.procedure_name)
            menu_summary[menu]["forms"].add(call.form_name)
            menu_summary[menu]["call_count"] += 1

        # set을 list로 변환
        result = {}
        for menu, data in menu_summary.items():
            result[menu] = {
                "procedures": sorted(list(data["procedures"])),
                "procedure_count": len(data["procedures"]),
                "forms": sorted(list(data["forms"])),
                "call_count": data["call_count"]
            }

        return dict(sorted(result.items()))

    def save_csv(self, output_path: str):
        """CSV 파일 저장"""
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # 헤더
            writer.writerow(["폼명", "프로시저명", "파일경로", "라인번호", "호출패턴"])

            # 데이터 (폼명으로 정렬)
            sorted_calls = sorted(self.procedure_calls, key=lambda x: (x.form_name, x.procedure_name))
            for call in sorted_calls:
                writer.writerow([
                    call.form_name,
                    call.procedure_name,
                    call.file_path,
                    call.line_number,
                    call.call_pattern
                ])

        print(f"CSV 저장 완료: {output_file}")

    def save_excel(self, output_path: str):
        """Excel 파일 저장 (3개 시트)"""
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # 스타일 정의
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === Sheet 1: 전체 목록 ===
        ws1 = wb.active
        ws1.title = "전체목록"

        headers1 = ["폼명", "프로시저명", "파일경로", "라인번호", "호출패턴"]
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        sorted_calls = sorted(self.procedure_calls, key=lambda x: (x.form_name, x.procedure_name))
        for row_idx, call in enumerate(sorted_calls, 2):
            ws1.cell(row=row_idx, column=1, value=call.form_name).border = thin_border
            ws1.cell(row=row_idx, column=2, value=call.procedure_name).border = thin_border
            ws1.cell(row=row_idx, column=3, value=call.file_path).border = thin_border
            ws1.cell(row=row_idx, column=4, value=call.line_number).border = thin_border
            ws1.cell(row=row_idx, column=5, value=call.call_pattern).border = thin_border

        self._auto_adjust_width(ws1)

        # === Sheet 2: 폼별 요약 ===
        ws2 = wb.create_sheet(title="폼별요약")

        headers2 = ["폼명", "프로시저수", "프로시저목록"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        sorted_forms = sorted(self.form_summaries.values(), key=lambda x: -x.procedure_count)
        for row_idx, summary in enumerate(sorted_forms, 2):
            ws2.cell(row=row_idx, column=1, value=summary.form_name).border = thin_border
            ws2.cell(row=row_idx, column=2, value=summary.procedure_count).border = thin_border
            # 프로시저 목록 (최대 10개 표시)
            proc_list = ", ".join(summary.procedures[:10])
            if len(summary.procedures) > 10:
                proc_list += f" ... (+{len(summary.procedures) - 10}개)"
            ws2.cell(row=row_idx, column=3, value=proc_list).border = thin_border

        self._auto_adjust_width(ws2)

        # === Sheet 3: 프로시저별 요약 ===
        ws3 = wb.create_sheet(title="프로시저별요약")

        headers3 = ["프로시저명", "호출횟수", "호출폼목록"]
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        sorted_procs = sorted(self.procedure_summaries.values(), key=lambda x: -x.call_count)
        for row_idx, summary in enumerate(sorted_procs, 2):
            ws3.cell(row=row_idx, column=1, value=summary.procedure_name).border = thin_border
            ws3.cell(row=row_idx, column=2, value=summary.call_count).border = thin_border
            ws3.cell(row=row_idx, column=3, value=", ".join(summary.calling_forms)).border = thin_border

        self._auto_adjust_width(ws3)

        # === Sheet 4: 메뉴별 요약 (NEW!) ===
        ws4 = wb.create_sheet(title="메뉴별요약")

        menu_summary = self.get_menu_based_summary()

        headers4 = ["메뉴", "프로시저수", "호출횟수", "관련폼"]
        for col, header in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, (menu, data) in enumerate(menu_summary.items(), 2):
            ws4.cell(row=row_idx, column=1, value=menu).border = thin_border
            ws4.cell(row=row_idx, column=2, value=data["procedure_count"]).border = thin_border
            ws4.cell(row=row_idx, column=3, value=data["call_count"]).border = thin_border
            ws4.cell(row=row_idx, column=4, value=", ".join(data["forms"][:5])).border = thin_border

        self._auto_adjust_width(ws4)

        # === Sheet 5: 메뉴별 상세 (NEW!) ===
        ws5 = wb.create_sheet(title="메뉴별상세")

        headers5 = ["메뉴", "프로시저명", "관련폼"]
        for col, header in enumerate(headers5, 1):
            cell = ws5.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        row_idx = 2
        for menu, data in menu_summary.items():
            for proc in data["procedures"]:
                # 이 프로시저를 호출하는 폼 목록
                calling_forms = self.procedure_summaries.get(proc, ProcedureSummary(proc)).calling_forms
                ws5.cell(row=row_idx, column=1, value=menu).border = thin_border
                ws5.cell(row=row_idx, column=2, value=proc).border = thin_border
                ws5.cell(row=row_idx, column=3, value=", ".join(calling_forms[:3])).border = thin_border
                row_idx += 1

        self._auto_adjust_width(ws5)

        # === Sheet 6: 통계 ===
        ws6 = wb.create_sheet(title="통계")

        stats = [
            ("총 프로시저 호출 수", len(self.procedure_calls)),
            ("고유 프로시저 수", len(self._get_unique_procedures())),
            ("관련 폼/클래스 수", len(self.form_summaries)),
            ("", ""),
            ("가장 많이 호출되는 프로시저 Top 10", ""),
        ]

        for row_idx, (label, value) in enumerate(stats, 1):
            ws6.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws6.cell(row=row_idx, column=2, value=value)

        # Top 10 프로시저
        top_procs = sorted(self.procedure_summaries.values(), key=lambda x: -x.call_count)[:10]
        for i, proc in enumerate(top_procs, 6):
            ws6.cell(row=i, column=1, value=proc.procedure_name)
            ws6.cell(row=i, column=2, value=proc.call_count)

        self._auto_adjust_width(ws6)

        wb.save(output_file)
        print(f"Excel 저장 완료: {output_file}")

    def _auto_adjust_width(self, ws):
        """컬럼 너비 자동 조정"""
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for row_idx in range(1, min(ws.max_row + 1, 100)):  # 최대 100행만 체크
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    max_length = max(max_length, length)

            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def print_statistics(self):
        """통계 출력"""
        print("\n" + "=" * 60)
        print("BOGUN2018 프로시저 매핑 통계")
        print("=" * 60)
        print(f"총 프로시저 호출 수: {len(self.procedure_calls)}")
        print(f"고유 프로시저 수: {len(self._get_unique_procedures())}")
        print(f"관련 폼/클래스 수: {len(self.form_summaries)}")

        # 메뉴별 요약
        menu_summary = self.get_menu_based_summary()
        print("\n" + "-" * 60)
        print("메뉴별 프로시저 분포")
        print("-" * 60)
        for menu, data in menu_summary.items():
            print(f"  {menu}: {data['procedure_count']}개 프로시저, {data['call_count']}회 호출")

        print("\n--- 가장 많이 호출되는 프로시저 Top 10 ---")
        sorted_procs = sorted(self.procedure_summaries.values(), key=lambda x: -x.call_count)[:10]
        for summary in sorted_procs:
            print(f"  {summary.procedure_name}: {summary.call_count}회")


def main():
    """메인 실행 함수"""
    # 경로 설정
    base_path = Path(__file__).parent.parent
    output_dir = base_path / "output"

    # 전체 BOGUN2018 폴더 스캔 (하위 디렉토리 모두 포함)
    bogun_root = base_path / "BOGUN2018"

    # 추출기 실행
    extractor = BogunProcedureExtractor(str(bogun_root))
    extractor.extract_all()

    # 결과 저장
    extractor.save_csv(str(output_dir / "bogun_procedure_mapping.csv"))
    extractor.save_excel(str(output_dir / "bogun_procedure_mapping.xlsx"))

    # 통계 출력
    extractor.print_statistics()


if __name__ == "__main__":
    main()
